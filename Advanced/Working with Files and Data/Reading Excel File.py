import openpyxl

def read_excel_file(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active  # Read the first sheet

        rows = sheet.max_row
        cols = sheet.max_column

        print(f"The Excel file has {rows} rows and {cols} columns.")
    except FileNotFoundError:
        print(f"Error: The file at '{file_path}' was not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

excel_file_path = "D:\\101 Python Course 2025\\Advanced\\Working with Files and Datas\\test.xlsx"
read_excel_file(excel_file_path)
